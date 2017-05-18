#! usr/bin/env python3
# SCOPUSAbstractMine.py - Takes the abstracts from a SCOPUS search csv output
# and writes each into its own text file for future mining

import openpyxl, os, re

wb = openpyxl.load_workbook('SCOPUSCLEAN.xlsx')
sheet = wb.get_sheet_by_name('scopus')

file_name = input("Enter base name for all files: ")
new_folder = file_name + '_Abstract_Files'
os.makedirs(new_folder)

whitelist = set('abcdefghijklmnopqrstuvwxyz-_ ')

#Dictionary to find terms that need to be combined with underscores
repldict = {'ssb':'sugarsweetenedbeverage',
	'ssbs':'sugarsweetenedbeverage',
	'sugar sweetened beverage':'sugarsweetenedbeverage',
	'sugar sweetened beverages':'sugarsweetenedbeverage',
	'sugarsweetened beverage':'sugarsweetenedbeverage',
	'sugarsweetened beverages':'sugarsweetenedbeverage',
	'sugarsweetenedbeverage':'sugarsweetenedbeverage',
	'sugarsweetenedbeverages':'sugarsweetenedbeverage',
	'sugar sweeten beverage' : 'sugarsweetenedbeverage',
	'sugar sweeten beve rage' : 'sugarsweetenedbeverage',
	'sugar sweeten beverages' : 'sugarsweetenedbeverage',
	'sugar-sweetened beverage' : 'sugarsweetenedbeverage',
	'sugar-sweetened beverages' : 'sugarsweetenedbeverage',
	'newyork':'newyork',
	'new york':'newyork',
	'new york city':'newyorkcity',
	'newyork city':'newyorkcity',
	'newyorkcity':'newyorkcity',
	'-':' '
	}

def replfunc(match):
	return repldict[match.group(0)]

# Creates regex based on dictionary
ssbregex = re.compile('|'.join(re.escape(x) for x in repldict))

for row in range (2, sheet.max_row + 1):
	docid = sheet['A' + str(row)].value
	abstract = sheet['Q' + str(row)].value
	sep = '©'
	abstract = abstract.split(sep, 1)[0]
	abstract = abstract.lower()
	abstract = ''.join(filter(whitelist.__contains__, abstract))
	abstract = ssbregex.sub(replfunc, abstract)

	# print('Writing results...')

	f_write = open(os.path.join(new_folder, file_name + '_' + str(docid) + '.txt'), 'w')
	f_write.write(abstract)
	f_write.close()

print('Done creating ' + str(sheet.max_row) + ' abstract files.')
